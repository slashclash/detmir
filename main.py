from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.common.proxy import Proxy, ProxyType
import time
from selenium.common.exceptions import NoSuchElementException, StaleElementReferenceException
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
import csv

def save_to_excel(products):
    """
    Save Products to XLSX file.
    :param products: list of products.
    :return: create XLSX file.
    """

    wb = Workbook()


    ws = wb.create_sheet("Данные")
    ws['A1'] = 'id'
    ws['B1'] = 'title'
    ws['C1'] = 'price'
    ws['D1'] = 'town'
    ws['E1'] = 'promo_price'
    ws['F1'] = 'url'


    medium_border = (Border(left=Side(style='medium'),
                            right=Side(style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(style='medium')))

    thin_border = (Border(left=Side(style='thin'),
                          right=Side(style='thin'),
                          top=Side(style='thin'),
                          bottom=Side(style='thin')))

    headers = ws[1]
    for cell in headers:
        cell.font = Font(bold=True)
        cell.border = medium_border

    for i, product in enumerate(products):
        _ = ws.cell(row=i+2, column=1, value=product[0])
        _ = ws.cell(row=i+2, column=2, value=product[1])
        _ = ws.cell(row=i+2, column=3, value=product[2])
        _ = ws.cell(row=i+2, column=4, value=product[3])
        _ = ws.cell(row=i+2, column=5, value=product[4])
        _ = ws.cell(row=i+2, column=6, value=product[5])

    for row in ws.iter_rows(min_row=2, max_col=3):
        for cell in row:
            cell.border = thin_border

    wb.remove(wb['Sheet'])
    file_name = "Data.xlsx"
    wb.save(file_name)
    print("Данные сохранены в файле {}".format(file_name))

def save_to_csv(products):
    file_name = "Data.csv"
    with open(file_name, 'w', newline='', encoding='utf-8') as csvfile:
        fieldnames = ['id', 'title', 'price', 'town', 'promo_price', 'url']
        writer = csv.writer(csvfile)
        writer.writerow(fieldnames)
        writer.writerows(products)
    print("Данные сохранены в файле {}".format(file_name))

PRODUCTS = []

PROXY = "127.0.0.1:9150"
proxy = Proxy({
    'proxyType': ProxyType.MANUAL,
    'httpProxy': PROXY,
    'ftpProxy': PROXY,
    'sslProxy': PROXY,
    'noProxy': '' # set this value as desired
    })

service = Service(r"C:\Users\Alex\PycharmProjects\geckodriver.exe")

driver = webdriver.Firefox(service=service, proxy=proxy)

time.sleep(5)
def choose_location(driver, index):
    location_button = driver.find_element(By.CLASS_NAME, "wI")
    location_button.click()
    location_towns = driver.find_elements(By.CLASS_NAME, "Js")
    location_towns[index].click()

def get_content(driver):
        town = driver.find_element(By.CLASS_NAME, "i_8").find_element(By.CLASS_NAME, "wJ").text
        print(town)
        print("Пролистываем все страницы с товарами.")
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(1)
        driver.execute_script("window.scroll(0,0)", "")
        button_find = 1
        while button_find == 1:
            try:
                foot = driver.find_element(By.CLASS_NAME, "ck").find_element(By.CSS_SELECTOR, "button")
                y_offset = foot.location['y']-50
                driver.execute_script("arguments[0].scrollIntoView();", foot)
                driver.execute_script("window.scroll(0, {})".format(y_offset), "")
                foot.click()
                driver.execute_script("window.scroll(0,0)", "")
            except (NoSuchElementException, StaleElementReferenceException) as e:
                button_find = 0

        driver.execute_script("window.scroll(0,0)", "")
        cards = driver.find_elements(By.CLASS_NAME, "M_7")
        print("Найдено товаров:", len(cards))
        try:
            for card in cards:
                i = 0
                url = card.get_attribute("href")
                item_id = url.split("/")[-2]
                title = card.find_element(By.CLASS_NAME, 'Nw').find_element(By.CLASS_NAME, "Na").text
                try:
                    prices = card.find_element(By.CLASS_NAME, "Nk").find_elements(By.CSS_SELECTOR, "p")

                    if len(prices) == 1:
                        promo_price = ""
                        price = prices[0].text[:-2]
                    else:
                        promo_price = prices[0].text[:-2]
                        price = prices[1].text[:-2]

                except NoSuchElementException:
                    promo_price = ""
                    price = "Нет в наличии"

                info = [item_id, title, price, town, promo_price, url]
                PRODUCTS.append(info)

        except Exception:
            pass

if __name__ == "__main__":
    driver.get("https://www.detmir.ru/catalog/index/name/lego/")
    time.sleep(1)
    choose_location(driver, 0)
    get_content(driver)
    driver.get("https://www.detmir.ru/catalog/index/name/lego/")
    time.sleep(1)
    choose_location(driver, 1)
    get_content(driver)
    save_to_excel(PRODUCTS)
    save_to_csv(PRODUCTS)
    driver.quit()
